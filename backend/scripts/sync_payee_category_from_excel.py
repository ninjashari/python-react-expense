"""
Sync payee and category fields from an Excel export to the database.

Usage:
    cd backend
    python scripts/sync_payee_category_from_excel.py --email user@example.com
    python scripts/sync_payee_category_from_excel.py --email user@example.com --dry-run

Match key: date + amount + description (case-insensitive strip) + account name
Strategy : always overwrite payee/category; create missing payees/categories automatically
Safe     : idempotent — re-running produces the same result
"""

import sys
import os
import argparse
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import date, datetime

# Allow imports from backend/
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import difflib

import openpyxl
from sqlalchemy import func
from sqlalchemy.orm import Session

FUZZY_THRESHOLD = 0.95  # minimum similarity ratio for fallback fuzzy match


def fuzzy_match(excel_desc: str, candidates: list) -> object | None:
    """Return the single candidate whose description best matches excel_desc at >= FUZZY_THRESHOLD, or None."""
    excel_norm = excel_desc.lower().strip()
    hits = []
    for tx in candidates:
        db_norm = (tx.description or "").lower().strip()
        ratio = difflib.SequenceMatcher(None, excel_norm, db_norm).ratio()
        if ratio >= FUZZY_THRESHOLD:
            hits.append((ratio, tx))
    if len(hits) == 1:
        return hits[0][1], hits[0][0]
    return None, None

# Log file setup
_log_file = None

def _log(msg: str):
    if _log_file:
        _log_file.write(msg + "\n")
        _log_file.flush()


def _open_log() -> tuple[object, str]:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(script_dir, f"sync_payee_category_{ts}.log")
    return open(path, "w", encoding="utf-8"), path

from database import SessionLocal
from models.users import User
from models.accounts import Account
from models.transactions import Transaction
from models.categories import Category
from models.payees import Payee
from utils.slug import create_slug

EXCEL_PATH = r"C:\Users\abhin\Downloads\all_transactions_export_website.xlsx"
BATCH_SIZE = 500


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return None


def to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, (datetime,)):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def clean_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_or_create_category(db: Session, user_id, name: str, dry_run: bool, created: list) -> object | None:
    if not name:
        return None
    slug = create_slug(name)
    cat = db.query(Category).filter_by(user_id=user_id, slug=slug).first()
    if not cat:
        if dry_run:
            created.append(("category", name))
            return None  # can't create in dry-run; caller handles None
        cat = Category(user_id=user_id, name=name, slug=slug, color="#6366f1")
        db.add(cat)
        db.flush()
        created.append(("category", name))
    return cat


def get_or_create_payee(db: Session, user_id, name: str, dry_run: bool, created: list) -> object | None:
    if not name:
        return None
    slug = create_slug(name)
    payee = db.query(Payee).filter_by(user_id=user_id, slug=slug).first()
    if not payee:
        if dry_run:
            created.append(("payee", name))
            return None
        payee = Payee(user_id=user_id, name=name, slug=slug, color="#6366f1")
        db.add(payee)
        db.flush()
        created.append(("payee", name))
    return payee


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(email: str, dry_run: bool):
    global _log_file
    _log_file, log_path = _open_log()
    db: Session = SessionLocal()
    try:
        # -- Resolve user
        user = db.query(User).filter(func.lower(User.email) == email.lower()).first()
        if not user:
            print(f"ERROR: No user found with email '{email}'")
            sys.exit(1)
        user_id = user.id
        print(f"User: {user.email} ({user_id})")

        # -- Build account map (name → id)
        accounts = db.query(Account).filter(Account.user_id == user_id).all()
        account_map = {a.name.lower(): a.id for a in accounts}
        print(f"Accounts loaded: {len(account_map)}")

        # -- Load Excel
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active

        headers = [clean_str(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}

        required = {"date", "description", "amount", "account", "category", "payee"}
        missing = required - set(col.keys())
        if missing:
            print(f"ERROR: Excel missing columns: {missing}")
            sys.exit(1)

        # -- Counters
        total = matched = not_found = ambiguous = skipped = errors = fuzzy_matched = 0
        new_created: list[tuple[str, str]] = []

        rows_since_commit = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total += 1
            try:
                raw_date = row[col["date"]]
                raw_desc = row[col["description"]]
                raw_amount = row[col["amount"]]
                raw_account = row[col["account"]]
                raw_category = row[col["category"]]
                raw_payee = row[col["payee"]]

                cat_name = clean_str(raw_category)
                payee_name = clean_str(raw_payee)

                # Skip if nothing to update
                if not cat_name and not payee_name:
                    skipped += 1
                    continue

                tx_date = to_date(raw_date)
                tx_amount = to_decimal(raw_amount)
                tx_desc = clean_str(raw_desc)
                account_key = clean_str(raw_account).lower()

                if tx_date is None or tx_amount is None:
                    _log(f"[ROW {row_idx}] SKIP: unparseable date/amount | date={raw_date!r} amount={raw_amount!r}")
                    skipped += 1
                    continue

                account_id = account_map.get(account_key)
                if account_id is None:
                    _log(f"[ROW {row_idx}] SKIP: unknown account | {raw_account!r}")
                    skipped += 1
                    continue

                # -- Phase 1: exact match (case-insensitive, trimmed)
                matches = (
                    db.query(Transaction)
                    .filter(
                        Transaction.user_id == user_id,
                        Transaction.account_id == account_id,
                        Transaction.date == tx_date,
                        Transaction.amount == tx_amount,
                        func.lower(func.trim(Transaction.description)) == tx_desc.lower(),
                    )
                    .all()
                )

                fuzzy_used = False

                if len(matches) == 0:
                    # -- Phase 2: fuzzy fallback — same date+amount, pick best description match
                    candidates = (
                        db.query(Transaction)
                        .filter(
                            Transaction.user_id == user_id,
                            Transaction.account_id == account_id,
                            Transaction.date == tx_date,
                            Transaction.amount == tx_amount,
                        )
                        .all()
                    )
                    tx_fuzzy, ratio = fuzzy_match(tx_desc, candidates)
                    if tx_fuzzy:
                        matches = [tx_fuzzy]
                        fuzzy_used = True
                        _log(
                            f"[ROW {row_idx}] FUZZY_MATCH (ratio={ratio:.3f}) | {tx_date} | {tx_amount} "
                            f"| excel={tx_desc[:60]!r} | db={tx_fuzzy.description[:60]!r}"
                        )

                if len(matches) == 0:
                    not_found += 1
                    _log(f"[ROW {row_idx}] NOT_FOUND | {tx_date} | {tx_amount} | {tx_desc[:80]!r} | {raw_account}")
                    continue

                if len(matches) > 1:
                    ambiguous += 1
                    _log(f"[ROW {row_idx}] AMBIGUOUS ({len(matches)} hits) | {tx_date} | {tx_amount} | {tx_desc[:80]!r} | {raw_account}")
                    continue

                tx = matches[0]

                # -- Resolve / auto-create category & payee
                cat = get_or_create_category(db, user_id, cat_name, dry_run, new_created)
                payee = get_or_create_payee(db, user_id, payee_name, dry_run, new_created)

                if not dry_run:
                    tx.category_id = cat.id if cat else None
                    tx.payee_id = payee.id if payee else None

                matched += 1
                if fuzzy_used:
                    fuzzy_matched += 1
                rows_since_commit += 1
                tag = "MATCHED_FUZZY" if fuzzy_used else "MATCHED"
                _log(f"[ROW {row_idx}] {tag} | {tx_date} | {tx_amount} | {tx_desc[:80]!r} | cat={cat_name!r} payee={payee_name!r}")

                if not dry_run and rows_since_commit >= BATCH_SIZE:
                    db.commit()
                    rows_since_commit = 0

            except Exception as exc:
                errors += 1
                _log(f"[ROW {row_idx}] ERROR | {exc}")
                db.rollback()
                rows_since_commit = 0

        if not dry_run and rows_since_commit > 0:
            db.commit()

        wb.close()

        # -- Summary
        new_cats = sum(1 for t, _ in new_created if t == "category")
        new_payees = sum(1 for t, _ in new_created if t == "payee")

        print()
        print("=" * 50)
        print(f"{'DRY RUN — no changes written' if dry_run else 'DONE'}")
        print("=" * 50)
        print(f"Total rows processed   : {total}")
        print(f"Matched & updated      : {matched}  (of which fuzzy: {fuzzy_matched})")
        print(f"Not found              : {not_found}")
        print(f"Ambiguous (skipped)    : {ambiguous}")
        print(f"Skipped (no data/err)  : {skipped}")
        print(f"New categories created : {new_cats}")
        print(f"New payees created     : {new_payees}")
        print(f"Errors                 : {errors}")
        print(f"Log file               : {log_path}")

        if new_created:
            print("\nNew items:")
            for kind, name in sorted(set(new_created)):
                print(f"  [{kind}] {name}")

    finally:
        db.close()
        if _log_file:
            _log_file.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sync payee/category from Excel export to DB")
    parser.add_argument("--email", required=True, help="User email to scope updates")
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without writing to DB")
    args = parser.parse_args()
    run(email=args.email, dry_run=args.dry_run)
