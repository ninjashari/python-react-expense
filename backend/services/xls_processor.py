import io
import pandas as pd
from typing import Tuple, Optional, Dict, Any
from openpyxl import load_workbook
from fastapi import HTTPException


class XLSProcessor:
    """Service for processing XLS/XLSX files and extracting text content"""
    
    def __init__(self):
        self.supported_extensions = ['.xls', '.xlsx']
    
    def process_xls_file(self, file_bytes: bytes, filename: str = "") -> Tuple[str, str]:
        """
        Process XLS/XLSX file and extract all text content
        
        Args:
            file_bytes: Raw bytes of the Excel file
            filename: Original filename for context
            
        Returns:
            Tuple of (extracted_text, processing_method)
        """
        try:
            # Determine file type
            if filename.lower().endswith('.xlsx') or self._is_xlsx_format(file_bytes):
                return self._process_xlsx(file_bytes)
            else:
                return self._process_xls(file_bytes)
                
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Failed to process Excel file: {str(e)}"
            )
    
    def _is_xlsx_format(self, file_bytes: bytes) -> bool:
        """Check if the file is XLSX format by examining magic bytes"""
        # XLSX files start with PK (ZIP magic number)
        return file_bytes[:2] == b'PK'
    
    def _process_xlsx(self, file_bytes: bytes) -> Tuple[str, str]:
        """Process XLSX file using openpyxl"""
        try:
            # Load workbook from bytes
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
            extracted_text = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                extracted_text.append(f"=== SHEET: {sheet_name} ===")
                
                # Extract all cell values
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to strings
                    row_values = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                    if row_values:  # Only add non-empty rows
                        extracted_text.append(" | ".join(row_values))
            
            return "\n".join(extracted_text), "openpyxl_xlsx"
            
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Failed to process XLSX file with openpyxl: {str(e)}"
            )
    
    def _process_xls(self, file_bytes: bytes) -> Tuple[str, str]:
        """Process XLS file using pandas"""
        try:
            # Read XLS file with pandas
            excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
            extracted_text = []
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                extracted_text.append(f"=== SHEET: {sheet_name} ===")
                
                # Convert DataFrame to text format
                # Include column headers
                if not df.empty:
                    headers = " | ".join(str(col) for col in df.columns if str(col) != 'nan')
                    if headers:
                        extracted_text.append(headers)
                    
                    # Add each row
                    for _, row in df.iterrows():
                        row_values = [str(val) for val in row.values if pd.notna(val) and str(val).strip()]
                        if row_values:
                            extracted_text.append(" | ".join(row_values))
            
            return "\n".join(extracted_text), "pandas_xls"
            
        except Exception as e:
            # Fallback to XLSX processing if XLS fails
            try:
                return self._process_xlsx(file_bytes)
            except:
                raise HTTPException(
                    status_code=400,
                    detail=f"Failed to process XLS file with both pandas and openpyxl: {str(e)}"
                )
    
    def validate_extracted_text(self, text: str) -> bool:
        """
        Validate if extracted text contains potential financial data
        
        Args:
            text: Extracted text content
            
        Returns:
            bool: True if text appears to contain financial data
        """
        if not text or len(text.strip()) < 50:
            return False
        
        text_upper = text.upper()
        
        # Look for financial indicators
        financial_indicators = [
            'TRANSACTION', 'DEBIT', 'CREDIT', 'BALANCE', 'AMOUNT',
            'DATE', 'DESCRIPTION', 'WITHDRAWAL', 'DEPOSIT',
            'ACCOUNT', 'BANK', 'STATEMENT', 'PAYMENT',
            'TRANSFER', 'ATM', 'POS', 'UPI', 'NEFT', 'IMPS',
            'CHEQUE', 'CASH', 'CHARGE', 'FEE', 'INTEREST'
        ]
        
        financial_score = sum(1 for indicator in financial_indicators if indicator in text_upper)
        
        # Look for date patterns (common in bank statements)
        import re
        date_patterns = len(re.findall(r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b', text))
        
        # Look for amount patterns (money values)
        amount_patterns = len(re.findall(r'\b\d{1,3}(?:,\d{3})*\.?\d{0,2}\b', text))
        
        # Scoring logic
        has_financial_keywords = financial_score >= 3
        has_date_patterns = date_patterns >= 5
        has_amount_patterns = amount_patterns >= 5
        
        return has_financial_keywords and (has_date_patterns or has_amount_patterns)
    
    def get_file_info(self, file_bytes: bytes, filename: str = "") -> Dict[str, Any]:
        """
        Get basic information about the Excel file
        
        Args:
            file_bytes: Raw bytes of the Excel file
            filename: Original filename
            
        Returns:
            Dict with file information
        """
        try:
            file_info = {
                "filename": filename,
                "size_bytes": len(file_bytes),
                "format": "xlsx" if self._is_xlsx_format(file_bytes) else "xls",
                "sheets": []
            }
            
            # Try to get sheet information
            try:
                if file_info["format"] == "xlsx":
                    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        # Count non-empty rows
                        row_count = sum(1 for row in sheet.iter_rows() if any(cell.value for cell in row))
                        file_info["sheets"].append({
                            "name": sheet_name,
                            "rows": row_count,
                            "max_row": sheet.max_row,
                            "max_column": sheet.max_column
                        })
                else:
                    excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        file_info["sheets"].append({
                            "name": sheet_name,
                            "rows": len(df),
                            "columns": len(df.columns)
                        })
            except:
                # If we can't get sheet info, that's okay
                pass
            
            return file_info
            
        except Exception as e:
            return {
                "filename": filename,
                "size_bytes": len(file_bytes),
                "format": "unknown",
                "error": str(e)
            }